import requests
import time
from bs4 import BeautifulSoup
import random
import csv
import json
import re
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from requests.adapters import HTTPAdapter, Retry

base_url = 'https://nextbillion.net/jobs/?jobs-page=1'
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

session = requests.Session()
session.headers.update(headers)

retry_strategy = Retry(
    total=5,
    backoff_factor=1,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["HEAD", "GET", "OPTIONS"]
)
adapter = HTTPAdapter(max_retries=retry_strategy)
session.mount("https://", adapter)
session.mount("http://", adapter)

def get_data(url):
    try:
        page = session.get(url, timeout=10)
        page.raise_for_status()
        time.sleep(random.uniform(2, 4))
        return BeautifulSoup(page.text, 'lxml')
    except requests.exceptions.RequestException as e:
        print(f'request failed: {e}')
        return None


def extract_application_info(text):
    emails = re.findall(r'[\w\.-]+@[\w\.-]+', text)
    links = re.findall(r'https?://\S+', text)
    
    application_info = {
        "emails": ', '.join(emails) if emails else None,
        "links": ', '.join(links) if links else None,
        "type": "email" if emails else ("link" if links else "unknown")
    }
    return application_info


def extract_basic_info(div):
    description_tag = div.find('div', class_='excerpt')
    description = description_tag.get_text(separator=" ", strip=True) if description_tag else "N/A"
    
    job_title_tag = div.find('h3')
    job_title = job_title_tag.get_text(separator=" ", strip=True) if job_title_tag else "N/A"
    
    return {
        "title": job_title,
        "description": description
    }


def extract_details(details_link):
    details_page = get_data(details_link)
    if details_page:
        content = details_page.find('div', class_='content')
        if content:
            return content.get_text(separator=" ", strip=True)
    return "N/A"


def extract_job_metadata(div):
    company_name = "N/A"
    location = "N/A"
    job_type = "N/A"
    deadline = "N/A"
    tags = "N/A"
    
    elements = div.find_all('dl')
    for element in elements:
        dt = element.find('dt')
        dd = element.find('dd')
        
        if not dt or not dd:
            continue
        
        label = dt.text.strip().lower()
        value = dd.text.strip()
        
        if 'organization' in label:
            company_name = value
        elif 'location' in label:
            location = value
        elif 'job type' in label or 'job_type' in label or 'jobtype' in label:
            job_type = value
        elif 'deadline' in label:
            deadline = value
        elif 'tag' in label:
            found_tags = [t.text.strip() for t in dd.find_all('a')]
            if found_tags:
                tags = ', '.join(found_tags)
    
    return {
        "company": company_name,
        "location": location,
        "job_type": job_type,
        "deadline": deadline,
        "tags": tags
    }


def scrape_data(soup):
    data = []
    divs = soup.find_all('li', class_='clearfix')
    
    for div in divs:
        basic_info = extract_basic_info(div)
        job_title = basic_info["title"]
        description = basic_info["description"]
        
        metadata = extract_job_metadata(div)
        
        detailed_description = "N/A"
        application_info = {"emails": None, "links": None, "type": "unknown"}
        
        learn_more = div.find('div', class_='clearfix')
        if learn_more:
            learn_more_link = learn_more.find('a', class_='learnmore')
            if learn_more_link and learn_more_link.get('href'):
                details_link = urljoin(base_url, learn_more_link['href'])
                detailed_description = extract_details(details_link)
                application_info = extract_application_info(detailed_description)
        
        if application_info["type"] == "email":
            application_method = application_info["emails"]
        elif application_info["type"] == "link":
            application_method = application_info["links"]
        else:
            application_method = "N/A"
        
        data.append([
            job_title,
            description,
            metadata["company"],
            metadata["location"],
            metadata["job_type"],
            metadata["deadline"],
            metadata["tags"],
            detailed_description,
            application_method,
        ])
    
    return data


def next_page(soup):
    next_button = soup.find('li', class_='nextlink')
    if next_button:
        next_link = next_button.find('a')
        if next_link and next_link.get('href'):
            return next_link['href']
    return None


def save_to_csv(data):
    with open('nextbillion.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Job Title',
            'Description',
            'Organization',
            'Location',
            'Job Type',
            'Deadline',
            'Tags',
            'Detailed Description',
            'Application Method',
        ])
        writer.writerows(data)


def save_to_json(data):
    headers = [
        'Job Title',
        'Description',
        'Organization',
        'Location',
        'Job Type',
        'Deadline',
        'Tags',
        'Detailed Description',
        'Application Method',
    ]
    
    json_data = []
    for row in data:
        job_dict = {headers[i]: row[i] for i in range(len(headers))}
        json_data.append(job_dict)
    
    with open('nextbillion.json', 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)


def save_to_xlsx(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs Data"
    
    headers = [
        'Job Title',
        'Description',
        'Organization',
        'Location',
        'Job Type',
        'Deadline',
        'Tags',
        'Detailed Description',
        'Application Method',
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_num).column_letter
        for row in ws.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save('nextbillion.xlsx')


def main():
    url = base_url
    all_data = []
    n = 1
    while url:
        soup = get_data(url)
        if not soup:
            break

        page_data = scrape_data(soup)
        all_data.extend(page_data)

        next_href = next_page(soup)
        url = urljoin(url, next_href) if next_href else None

        print(f'scraping page {n}....')
        n += 1

    save_to_csv(all_data)
    save_to_xlsx(all_data)
    save_to_json(all_data)
    print('Data saved to nextbillion.csv, nextbillion.xlsx, and nextbillion.json')

if __name__ == '__main__' :
    main()



